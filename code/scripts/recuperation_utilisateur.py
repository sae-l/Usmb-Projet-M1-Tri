#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import ldap,string
from openpyxl import load_workbook
import unidecode 
# Initialise une connection au local pour faire des recherche
# Prend l'adresse du serveur, et les credential du l utilisateur sous lequel on veut faire la recherche
class Connection_ldap:

    def __init__(self,adresse,user,mdp):
        self.__connection=ldap.initialize("ldaps://%s:636"%(adresse), bytes_mode=False)
        self.__connection.simple_bind_s("uid=%s,ou=people,ou=uds,dc=agalan,dc=org"%(user), "%s"%(mdp))

    # Contacte le serveur ldap et recuperela liste des étudiant d'un groupe
    # Prend en entrée le groupe recherché
    def import_groupe_ldap(self,groupe):
        liste_finale=[]
        try:
            results = self.__connection.search_ext_s('dc=agalan,dc=org', ldap.SCOPE_SUBTREE, "cn=%s"%(groupe),['member'])
            results=results[0][1]["member"]
            for i in range(0,len(results)-1):
                dn = results[i].decode('utf8')
                if dn[:3]!="uid":
                    del results[i]
                else:
                    sp=dn.split(',')
                    liste_finale.append(sp[0][4:])
            return liste_finale
        except:
            print("Impossible de contacter le serveur")

    #Va chercher le nom et prenom d un utilisiteur sur le ldap
    #Prend en entrée l uid de la personne
    def import_user_ldap(self,pseudo):
        res=[]
        try:
            results = self.__connection.search_ext_s('dc=agalan,dc=org', ldap.SCOPE_SUBTREE, "uid=%s"%(pseudo),['sn','givenname'])
            tuples=(results[0][1]["sn"][0].decode('utf8'),results[0][1]["givenName"][0].decode('utf8'))
            return tuples
        except:
            print("Impossible de contacter le serveur")

    def lister_utilisateurs_groupe(self,groupe):
        liste_utilisateur = []
        liste_uid=self.import_groupe_ldap(groupe)
        for i in range(0,len(liste_uid)-1):
            tuple_user=self.import_user_ldap(liste_uid[i])
            liste_utilisateur.append([liste_uid[i],tuple_user[0],tuple_user[1]])
        return liste_utilisateur

#Lit les information d'un fichier xlsx et retourne le resultat sous la forme d'une liste
class Traitement_xlsx:
    __wb=None
    def __init__(self,fichier):
        self.__wb = load_workbook(filename='%s'%(fichier))       

    def traitement_fichier(self):
        liste=[]
        for sheet in self.__wb:
            for row in sheet.iter_rows(max_col=3,values_only=True):
                if row[0]!="Nom":
                    liste.append(row)
        return liste

#primitive : ("fichier xlsx","liste_ldap","serveur","user","mdp")
#Rend une liste avec les utilisateur present dans les 2 liste et une avec les utilisateur absent
def Recuperation_utilisateur(fichier,liste_groupe,serveur,user,mdp):
    liste_utilisateur = []
    liste_absent = []
    liste_ldap=[]
    liste_fichier=[]
    # recuperer adresse serveur/ credential
    conn_ldap=Connection_ldap(serveur,user,mdp)
    liste_ldap=conn_ldap.lister_utilisateurs_groupe(liste_groupe)
    del conn_ldap
    # print(liste_ldap)
    # Recupere la liste du fichier xlsx
    doc = Traitement_xlsx(fichier)
    liste_fichier=doc.traitement_fichier()
    del doc
    # print(liste_fichier)
    #liste ldap   [[uid, nom,prenom],[uid, nom,prenom],[uid, nom,prenom],...]
    #liste fichier   [[nom,prenom,groupe],[nom,prenom,groupe],...]
            
    #normalement la liste ldap est plus longue
    #longueur = len(self.__liste_fichier) if len(self.__liste_fichier) > len(self.__liste_ldap) else len(self.__liste_ldap)
    # On fait la liste des Utilisateurs present dans le ldap, on remplie tout les information si il sont dans les deux 
    # et si il ne sont que dans le ldap  on laisse vide
    for i in range(0,len(liste_ldap)):
        bool_match=False
        for j in range(0,len(liste_fichier)):
            if ((unidecode.unidecode(liste_ldap[i][1].upper()) == unidecode.unidecode(liste_fichier[j][0].upper())) and (unidecode.unidecode(liste_ldap[i][2].upper())== unidecode.unidecode(liste_fichier[j][1].upper()))):
                liste_utilisateur.append([liste_ldap[i][0],liste_ldap[i][1],liste_ldap[i][2],liste_fichier[j][2]])
                bool_match=True
            else:
                if (j==len(liste_fichier)-1 and bool_match==False):
                    liste_utilisateur.append([liste_ldap[i][0],liste_ldap[i][1],liste_ldap[i][2],""])
    # On passe dans le sens inverse pour les utilisateurs present que dans le fichier xlsx
    for k in range(0,len(liste_fichier)):
        match_bool = False
        for l in range(0,len(liste_utilisateur)):
            if (unidecode.unidecode(liste_utilisateur[l][1].upper()) == unidecode.unidecode(liste_fichier[k][0].upper())) and (unidecode.unidecode(liste_utilisateur[l][2].upper()) == unidecode.unidecode(liste_fichier[k][1].upper())):
                match_bool=True
            if (l==len(liste_ldap)-1 and match_bool==False):
                liste_utilisateur.append(["",liste_fichier[k][0],liste_fichier[k][1],liste_fichier[k][2]])
    return (liste_utilisateur)

# ls=Recuperation_utilisateur("doc.xlsx","etudiants-m1-sts-rt-tri","ldap-bourget.univ-smb.fr","user","mdp")
# print(ls)
# liste < tuple< dict< liste ... 
